from fastapi import APIRouter, Depends, HTTPException, Query
from typing import Optional, List
from datetime import datetime, date
from app.auth import get_current_user
from app.database import get_supabase_client
from app.models import UserResponse
import io
from fastapi.responses import StreamingResponse
import pytz

# Para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
except ImportError:
    pass

# Para PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    pass

router = APIRouter(prefix="/reportes", tags=["reportes"])


def verificar_admin(current_user: UserResponse):
    """Verifica que el usuario sea administrador o supervisor"""
    if current_user.rol not in ["admin", "administrador", "supervisor"]:
        raise HTTPException(
            status_code=403,
            detail="No tienes permisos para acceder a reportes"
        )


@router.get("/visitas")
async def obtener_reporte_visitas(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    usuario_id: Optional[str] = Query(None),
    servicio_id: Optional[int] = Query(None),  # OBLIGATORIO desde frontend
    punto_qr_id: Optional[str] = Query(None),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Obtiene reporte de visitas con filtros opcionales.
    servicio_id es obligatorio desde el frontend.
    """
    verificar_admin(current_user)
    
    supabase = get_supabase_client()
    
    try:
        # Construir query base
        query = supabase.table("visitas").select("*")
        
        # Aplicar filtros
        if fecha_inicio:
            query = query.gte("created_at", fecha_inicio)
        if fecha_fin:
            # Agregar un día para incluir todo el día final
            fecha_fin_dt = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
            from datetime import timedelta
            fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
            query = query.lt("created_at", fecha_fin_dt.isoformat())
        if usuario_id:
            query = query.eq("usuario_id", usuario_id)
        if punto_qr_id:
            query = query.eq("punto_qr_id", punto_qr_id)
        
        query = query.order("created_at", desc=True)
        
        response = query.execute()
        visitas = response.data
        
        # ⭐ FILTRAR POR SERVICIO_ID (desde puntos_qr) ⭐
        if servicio_id:
            # Obtener puntos del servicio
            puntos_resp = supabase.table("puntos_qr").select("id").eq("servicio_id", servicio_id).execute()
            puntos_ids = [p["id"] for p in puntos_resp.data]
            
            # Filtrar visitas que pertenecen a esos puntos
            visitas = [v for v in visitas if v.get("punto_qr_id") in puntos_ids]
        
        # Enriquecer datos con nombres (queries separadas)
        usuarios_cache = {}
        puntos_cache = {}
        
        for visita in visitas:
            # Obtener nombre de usuario/guardia
            if visita.get("guardia_id"):
                if visita["guardia_id"] not in usuarios_cache:
                    try:
                        user_resp = supabase.table("usuarios").select("nombre, email").eq("id", visita["guardia_id"]).execute()
                        if user_resp.data:
                            usuarios_cache[visita["guardia_id"]] = user_resp.data[0]
                    except:
                        usuarios_cache[visita["guardia_id"]] = {"nombre": "Usuario desconocido", "email": ""}
                
                visita["usuario_nombre"] = usuarios_cache[visita["guardia_id"]].get("nombre", "Desconocido")
                visita["usuario_email"] = usuarios_cache[visita["guardia_id"]].get("email", "")
            
            # Obtener nombre de punto QR
            if visita.get("punto_qr_id"):
                if visita["punto_qr_id"] not in puntos_cache:
                    try:
                        punto_resp = supabase.table("puntos_qr").select("nombre, qr_code").eq("id", visita["punto_qr_id"]).execute()
                        if punto_resp.data:
                            puntos_cache[visita["punto_qr_id"]] = punto_resp.data[0]
                    except:
                        puntos_cache[visita["punto_qr_id"]] = {"nombre": "Punto desconocido", "qr_code": ""}
                
                visita["punto_nombre"] = puntos_cache[visita["punto_qr_id"]].get("nombre", "Desconocido")
                visita["punto_codigo"] = puntos_cache[visita["punto_qr_id"]].get("qr_code", "")
            
            # Asegurar que observacion esté presente (puede ser None o vacío)
            if "observacion" not in visita or visita["observacion"] is None:
                visita["observacion"] = ""
        
        # Calcular estadísticas
        estadisticas = {
            "total_visitas": len(visitas),
            "usuarios_unicos": len(set(v.get("guardia_id") for v in visitas if v.get("guardia_id"))),
            "puntos_visitados": len(set(v.get("punto_qr_id") for v in visitas if v.get("punto_qr_id"))),
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin
        }
        
        return {
            "visitas": visitas,
            "estadisticas": estadisticas
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al obtener visitas: {str(e)}")


@router.get("/puntos-ranking")
async def obtener_ranking_puntos(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    servicio_id: Optional[int] = Query(None),
    limit: int = Query(10, ge=1, le=100),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Obtiene ranking de puntos QR más visitados
    """
    verificar_admin(current_user)
    
    supabase = get_supabase_client()
    
    try:
        # Obtener visitas con filtros
        query = supabase.table("visitas").select("punto_qr_id")
        
        if fecha_inicio:
            query = query.gte("created_at", fecha_inicio)
        if fecha_fin:
            fecha_fin_dt = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
            from datetime import timedelta
            fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
            query = query.lt("created_at", fecha_fin_dt.isoformat())
        
        response = query.execute()
        visitas = response.data
        
        # ⭐ FILTRAR POR SERVICIO_ID ⭐
        if servicio_id:
            puntos_resp = supabase.table("puntos_qr").select("id").eq("servicio_id", servicio_id).execute()
            puntos_ids = [p["id"] for p in puntos_resp.data]
            visitas = [v for v in visitas if v.get("punto_qr_id") in puntos_ids]
        
        # Contar visitas por punto
        contador = {}
        for visita in visitas:
            punto_id = visita.get("punto_qr_id")
            if punto_id:
                contador[punto_id] = contador.get(punto_id, 0) + 1
        
        # Ordenar por cantidad de visitas
        ranking = sorted(contador.items(), key=lambda x: x[1], reverse=True)[:limit]
        
        # Obtener información de los puntos
        resultado = []
        for punto_id, visitas_count in ranking:
            try:
                punto_resp = supabase.table("puntos_qr").select("*").eq("id", punto_id).execute()
                if punto_resp.data:
                    punto = punto_resp.data[0]
                    resultado.append({
                        "punto_id": punto_id,
                        "nombre": punto.get("nombre"),
                        "codigo": punto.get("qr_code"),
                        "total_visitas": visitas_count
                    })
            except:
                continue
        
        return {
            "ranking": resultado,
            "total_puntos": len(resultado)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener ranking: {str(e)}")


@router.get("/alertas")
async def obtener_reporte_alertas(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    tipo: Optional[str] = Query(None),
    leido: Optional[bool] = Query(None),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Obtiene reporte de alertas con filtros
    """
    verificar_admin(current_user)
    
    supabase = get_supabase_client()
    
    try:
        query = supabase.table("alertas").select("*")
        
        if fecha_inicio:
            query = query.gte("created_at", fecha_inicio)
        if fecha_fin:
            fecha_fin_dt = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
            from datetime import timedelta
            fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
            query = query.lt("created_at", fecha_fin_dt.isoformat())
        if tipo:
            query = query.eq("tipo", tipo)
        if leido is not None:
            query = query.eq("leido", leido)
        
        query = query.order("created_at", desc=True)
        
        response = query.execute()
        alertas = response.data
        
        # Enriquecer con nombres
        for alerta in alertas:
            if alerta.get("usuario_id"):
                try:
                    user_resp = supabase.table("usuarios").select("nombre").eq("id", alerta["usuario_id"]).execute()
                    if user_resp.data:
                        alerta["usuario_nombre"] = user_resp.data[0].get("nombre", "Desconocido")
                except:
                    alerta["usuario_nombre"] = "Desconocido"
        
        # Estadísticas
        estadisticas = {
            "total_alertas": len(alertas),
            "alertas_leidas": sum(1 for a in alertas if a.get("leido")),
            "alertas_pendientes": sum(1 for a in alertas if not a.get("leido"))
        }
        
        return {
            "alertas": alertas,
            "estadisticas": estadisticas
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener alertas: {str(e)}")
    
@router.post("/exportar-excel")
async def exportar_excel(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    servicio_id: Optional[int] = Query(None),
    tipo: str = Query("visitas", regex="^(visitas|ranking|alertas|incidencias)$"),
    timezone: str = Query("UTC"),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Exporta reportes a Excel
    """
    verificar_admin(current_user)
    
    supabase = get_supabase_client()
    
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Obtener zona horaria del usuario
        try:
            user_tz = pytz.timezone(timezone)
        except:
            user_tz = pytz.UTC
        
        if tipo == "visitas" or tipo == "incidencias":
            # Obtener datos de visitas
            query = supabase.table("visitas").select("*")
            if fecha_inicio:
                query = query.gte("created_at", fecha_inicio)
            if fecha_fin:
                fecha_fin_dt = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
                from datetime import timedelta
                fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
                query = query.lt("created_at", fecha_fin_dt.isoformat())
            
            # Filtrar por tipo de visita si es incidencias
            if tipo == "incidencias":
                query = query.eq("tipo", "incidencia")
            
            query = query.order("created_at", desc=True)
            response = query.execute()
            visitas = response.data
            
            # ⭐ FILTRAR POR SERVICIO_ID ⭐
            if servicio_id:
                puntos_resp = supabase.table("puntos_qr").select("id").eq("servicio_id", servicio_id).execute()
                puntos_ids = [p["id"] for p in puntos_resp.data]
                visitas = [v for v in visitas if v.get("punto_qr_id") in puntos_ids]
            
            # Título y metadatos
            ws.title = "Reporte de Incidencias" if tipo == "incidencias" else "Reporte de Visitas"
            ws['A1'] = f"REPORTE DE {'INCIDENCIAS' if tipo == 'incidencias' else 'VISITAS'} - ACRUX 360"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:F1')
            
            # Convertir hora actual a zona horaria del usuario
            now_utc = datetime.now(pytz.UTC)
            now_local = now_utc.astimezone(user_tz)
            ws['A2'] = f"Generado: {now_local.strftime('%d/%m/%Y %H:%M')}"
            
            if fecha_inicio or fecha_fin:
                ws['A3'] = f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actualidad'}"
            
            # Headers - COLUMNAS ACTUALIZADAS
            headers = ["Fecha", "Hora", "Punto Visitado", "Guardia", "Estatus", "Observaciones"]
            ws.append([])  # Fila vacía
            ws.append(headers)
            
            header_row = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Enriquecer y agregar datos
            usuarios_cache = {}
            puntos_cache = {}
            
            for visita in visitas:
                # Obtener nombres
                usuario_nombre = "Desconocido"
                if visita.get("guardia_id"):
                    if visita["guardia_id"] not in usuarios_cache:
                        try:
                            user_resp = supabase.table("usuarios").select("nombre, email").eq("id", visita["guardia_id"]).execute()
                            if user_resp.data:
                                usuarios_cache[visita["guardia_id"]] = user_resp.data[0]
                        except:
                            usuarios_cache[visita["guardia_id"]] = {"nombre": "Desconocido", "email": ""}
                    
                    usuario_nombre = usuarios_cache[visita["guardia_id"]].get("nombre", "Desconocido")
                
                punto_nombre = "Desconocido"
                if visita.get("punto_qr_id"):
                    if visita["punto_qr_id"] not in puntos_cache:
                        try:
                            punto_resp = supabase.table("puntos_qr").select("nombre, qr_code").eq("id", visita["punto_qr_id"]).execute()
                            if punto_resp.data:
                                puntos_cache[visita["punto_qr_id"]] = punto_resp.data[0]
                        except:
                            puntos_cache[visita["punto_qr_id"]] = {"nombre": "Desconocido", "qr_code": ""}
                    
                    punto_nombre = puntos_cache[visita["punto_qr_id"]].get("nombre", "Desconocido")
                
                # Formatear fecha con zona horaria del usuario
                created_at = visita.get("created_at", "")
                fecha_str = ""
                hora_str = ""
                if created_at:
                    try:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        dt_local = dt.astimezone(user_tz)
                        fecha_str = dt_local.strftime('%d/%m/%Y')
                        hora_str = dt_local.strftime('%H:%M')
                    except:
                        pass
                
                # COLUMNAS ACTUALIZADAS: Fecha, Hora, Punto, Guardia, Estatus, Observaciones
                row_data = [
                    fecha_str,
                    hora_str,
                    punto_nombre,
                    usuario_nombre,
                    "Visitado",  # Estatus
                    visita.get("observacion", "")
                ]
                ws.append(row_data)
                
                # Aplicar bordes
                for col in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col).border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12  # Fecha
            ws.column_dimensions['B'].width = 8   # Hora
            ws.column_dimensions['C'].width = 30  # Punto
            ws.column_dimensions['D'].width = 25  # Guardia
            ws.column_dimensions['E'].width = 12  # Estatus
            ws.column_dimensions['F'].width = 40  # Observaciones
            
            # Agregar hoja de estadísticas
            ws_stats = wb.create_sheet("Estadísticas")
            ws_stats['A1'] = "ESTADÍSTICAS DEL PERÍODO"
            ws_stats['A1'].font = Font(size=14, bold=True)
            ws_stats.append([])
            ws_stats.append(["Total de visitas:", len(visitas)])
            ws_stats.append(["Usuarios únicos:", len(set(v.get("guardia_id") for v in visitas if v.get("guardia_id")))])
            ws_stats.append(["Puntos visitados:", len(set(v.get("punto_qr_id") for v in visitas if v.get("punto_qr_id")))])
            
        elif tipo == "ranking":
            ws.title = "Ranking de Puntos"
            # Implementar similar a visitas
            
        elif tipo == "alertas":
            ws.title = "Reporte de Alertas"
            # Implementar similar a visitas
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar Excel: {str(e)}")


@router.post("/exportar-pdf")
async def exportar_pdf(
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    servicio_id: Optional[int] = Query(None),
    tipo: str = Query("visitas", regex="^(visitas|ranking|alertas|incidencias)$"),
    timezone: str = Query("UTC"),
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Exporta reportes a PDF
    """
    verificar_admin(current_user)
    
    supabase = get_supabase_client()
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Obtener zona horaria del usuario
        try:
            user_tz = pytz.timezone(timezone)
        except:
            user_tz = pytz.UTC
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        elements.append(Paragraph(f"REPORTE DE {tipo.upper()} - ACRUX 360", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadatos con zona horaria local
        meta_style = styles['Normal']
        now_utc = datetime.now(pytz.UTC)
        now_local = now_utc.astimezone(user_tz)
        elements.append(Paragraph(f"<b>Generado:</b> {now_local.strftime('%d/%m/%Y %H:%M')}", meta_style))
        
        if fecha_inicio or fecha_fin:
            elements.append(Paragraph(f"<b>Período:</b> {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actualidad'}", meta_style))
        elements.append(Spacer(1, 0.3*inch))
        
        if tipo == "visitas" or tipo == "incidencias":
            # Obtener datos
            query = supabase.table("visitas").select("*")
            if fecha_inicio:
                query = query.gte("created_at", fecha_inicio)
            if fecha_fin:
                fecha_fin_dt = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
                from datetime import timedelta
                fecha_fin_dt = fecha_fin_dt + timedelta(days=1)
                query = query.lt("created_at", fecha_fin_dt.isoformat())
            
            # Filtrar por tipo de visita si es incidencias
            if tipo == "incidencias":
                query = query.eq("tipo", "incidencia")
            
            query = query.order("created_at", desc=True).limit(100)  # Limitar para PDF
            response = query.execute()
            visitas = response.data
            
            # ⭐ FILTRAR POR SERVICIO_ID ⭐
            if servicio_id:
                puntos_resp = supabase.table("puntos_qr").select("id").eq("servicio_id", servicio_id).execute()
                puntos_ids = [p["id"] for p in puntos_resp.data]
                visitas = [v for v in visitas if v.get("punto_qr_id") in puntos_ids]
            
            # Estadísticas
            stats_data = [
                ["Estadística", "Valor"],
                ["Total de visitas", str(len(visitas))],
                ["Usuarios únicos", str(len(set(v.get("guardia_id") for v in visitas if v.get("guardia_id"))))],
                ["Puntos visitados", str(len(set(v.get("punto_qr_id") for v in visitas if v.get("punto_qr_id"))))]
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(stats_table)
            elements.append(Spacer(1, 0.5*inch))
            
            # Tabla de visitas (limitada a primeras 50) - COLUMNAS ACTUALIZADAS
            elements.append(Paragraph("<b>ÚLTIMAS VISITAS (Máx. 50)</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.2*inch))
            
            table_data = [["Fecha", "Hora", "Punto", "Guardia"]]
            
            usuarios_cache = {}
            puntos_cache = {}
            
            for visita in visitas[:50]:
                # Obtener nombres
                usuario_nombre = "Desconocido"
                if visita.get("guardia_id"):
                    if visita["guardia_id"] not in usuarios_cache:
                        try:
                            user_resp = supabase.table("usuarios").select("nombre").eq("id", visita["guardia_id"]).execute()
                            if user_resp.data:
                                usuarios_cache[visita["guardia_id"]] = user_resp.data[0].get("nombre", "Desconocido")
                        except:
                            usuarios_cache[visita["guardia_id"]] = "Desconocido"
                    usuario_nombre = usuarios_cache[visita["guardia_id"]]
                
                punto_nombre = "Desconocido"
                if visita.get("punto_qr_id"):
                    if visita["punto_qr_id"] not in puntos_cache:
                        try:
                            punto_resp = supabase.table("puntos_qr").select("nombre").eq("id", visita["punto_qr_id"]).execute()
                            if punto_resp.data:
                                puntos_cache[visita["punto_qr_id"]] = punto_resp.data[0].get("nombre", "Desconocido")
                        except:
                            puntos_cache[visita["punto_qr_id"]] = "Desconocido"
                    punto_nombre = puntos_cache[visita["punto_qr_id"]]
                
                # Formatear fecha con zona horaria del usuario
                created_at = visita.get("created_at", "")
                fecha_str = ""
                hora_str = ""
                if created_at:
                    try:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        dt_local = dt.astimezone(user_tz)
                        fecha_str = dt_local.strftime('%d/%m/%Y')
                        hora_str = dt_local.strftime('%H:%M')
                    except:
                        pass
                
                table_data.append([fecha_str, hora_str, punto_nombre, usuario_nombre])
            
            visits_table = Table(table_data, colWidths=[1.5*inch, 1*inch, 2.5*inch, 2*inch])
            visits_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(visits_table)
        
        # Footer
        elements.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Sistema de Recorridas QR - Acrux 360", footer_style))
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar PDF: {str(e)}")